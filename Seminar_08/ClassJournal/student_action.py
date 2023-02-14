import user_interface as ui
from openpyxl import load_workbook
def show_scores():
    name = ui.find_name()
    if name == '0':
        ui.goodbye()
    else:
        w_book = load_workbook('journal.xlsx')
        w_sheet = w_book[name]
        print('*\t*\t*\t*\t*\t*\t*')
        for i in range(1, w_sheet.max_row + 1):
            for j in range(1, w_sheet.max_column + 1):
                temp = w_sheet.cell(row=i, column=j)
                print(temp.value, end=' ')
            print()
        print('*\t*\t*\t*\t*\t*\t*')
        ui.goodbye()
